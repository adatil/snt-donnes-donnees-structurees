#!/usr/bin/env python3
import openpyxl
import pandas as pd

# Lire avec openpyxl pour voir la structure exacte
wb = openpyxl.load_workbook("Liste des films ayant réalisé plus d'un million d'entrées -.xlsx")

print("Noms des feuilles:")
print(wb.sheetnames)

# Lire chaque feuille
for sheet_name in wb.sheetnames[:5]:  # Limiter à 5 premières feuilles
    print(f"\n\n{'='*80}")
    print(f"FEUILLE: {sheet_name}")
    print('='*80)

    # Lire avec pandas
    try:
        df = pd.read_excel("Liste des films ayant réalisé plus d'un million d'entrées -.xlsx",
                          sheet_name=sheet_name, header=None)
        print(f"Dimensions: {df.shape}")
        print("\nPremières lignes:")
        print(df.head(15))

        # Si la feuille est pour une année spécifique (2024 ou 2023)
        if sheet_name in ['2024', '2023', '2022', '2021', '2020']:
            print(f"\n\nToutes les données de {sheet_name}:")
            print(df.to_string())
    except Exception as e:
        print(f"Erreur: {e}")

print("\n\nNombre total de feuilles:", len(wb.sheetnames))
